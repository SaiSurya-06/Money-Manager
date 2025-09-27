from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import io
import openpyxl
import logging
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from .models import Account, Transaction, RecurringTransaction, TransactionTag
from .forms import (
    AccountForm, TransactionForm, TransactionFilterForm,
    RecurringTransactionForm, BulkTransactionUploadForm
)
from moneymanager.apps.core.models import Category


class TransactionListView(LoginRequiredMixin, ListView):
    """List all transactions with filtering and pagination."""
    model = Transaction
    template_name = 'transactions/transaction_list.html'
    context_object_name = 'transactions'
    paginate_by = 25

    def get_queryset(self):
        queryset = Transaction.objects.filter(is_active=True).select_related(
            'category', 'account', 'user'
        )

        # Filter by family group or user
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(user=self.request.user, family_group__isnull=True)

        # Apply filters
        form = TransactionFilterForm(
            self.request.GET,
            user=self.request.user,
            family_group=getattr(self.request, 'current_family_group', None)
        )

        if form.is_valid():
            if form.cleaned_data.get('search'):
                search_term = form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(description__icontains=search_term) |
                    Q(notes__icontains=search_term)
                )

            if form.cleaned_data.get('transaction_type'):
                queryset = queryset.filter(
                    transaction_type=form.cleaned_data['transaction_type']
                )

            if form.cleaned_data.get('category'):
                queryset = queryset.filter(category=form.cleaned_data['category'])

            if form.cleaned_data.get('account'):
                queryset = queryset.filter(
                    Q(account=form.cleaned_data['account']) |
                    Q(to_account=form.cleaned_data['account']) |
                    Q(from_account=form.cleaned_data['account'])
                )

            # Date filtering
            period = form.cleaned_data.get('period')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')

            if period == 'today':
                queryset = queryset.filter(date=timezone.now().date())
            elif period == 'week':
                start_week = timezone.now().date() - timedelta(days=timezone.now().weekday())
                queryset = queryset.filter(date__gte=start_week)
            elif period == 'month':
                start_month = timezone.now().date().replace(day=1)
                queryset = queryset.filter(date__gte=start_month)
            elif period == 'quarter':
                month = timezone.now().month
                quarter_start = ((month - 1) // 3) * 3 + 1
                start_quarter = timezone.now().date().replace(month=quarter_start, day=1)
                queryset = queryset.filter(date__gte=start_quarter)
            elif period == 'year':
                start_year = timezone.now().date().replace(month=1, day=1)
                queryset = queryset.filter(date__gte=start_year)
            elif period == 'custom' and date_from and date_to:
                queryset = queryset.filter(date__range=[date_from, date_to])

        return queryset.order_by('-date', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = TransactionFilterForm(
            self.request.GET,
            user=self.request.user,
            family_group=getattr(self.request, 'current_family_group', None)
        )

        # Calculate summary stats
        queryset = self.get_queryset()
        context['total_income'] = queryset.filter(
            transaction_type='income'
        ).aggregate(total=Sum('amount'))['total'] or 0

        context['total_expenses'] = queryset.filter(
            transaction_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or 0

        context['net_amount'] = context['total_income'] - context['total_expenses']

        return context


class TransactionCreateView(LoginRequiredMixin, CreateView):
    """Create a new transaction."""
    model = Transaction
    form_class = TransactionForm
    template_name = 'transactions/transaction_form.html'
    success_url = reverse_lazy('transactions:list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Transaction created successfully!')
        return response


class TransactionUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing transaction."""
    model = Transaction
    form_class = TransactionForm
    template_name = 'transactions/transaction_form.html'
    success_url = reverse_lazy('transactions:list')

    def get_queryset(self):
        queryset = Transaction.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(user=self.request.user, family_group__isnull=True)
        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Transaction updated successfully!')
        return response


class TransactionDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a transaction (soft delete)."""
    model = Transaction
    template_name = 'transactions/transaction_confirm_delete.html'
    success_url = reverse_lazy('transactions:list')

    def get_queryset(self):
        queryset = Transaction.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(user=self.request.user, family_group__isnull=True)
        return queryset

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # Soft delete
        self.object.is_active = False
        self.object.save()
        messages.success(request, 'Transaction deleted successfully!')
        return HttpResponseRedirect(self.success_url)


class AccountListView(LoginRequiredMixin, ListView):
    """List all accounts."""
    model = Account
    template_name = 'transactions/account_list.html'
    context_object_name = 'accounts'

    def get_queryset(self):
        queryset = Account.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(owner=self.request.user, family_group__isnull=True)
        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        accounts = self.get_queryset().filter(include_in_totals=True)

        context['total_balance'] = sum(account.current_balance for account in accounts)
        context['account_types'] = {}

        for account in accounts:
            account_type = account.get_account_type_display()
            if account_type not in context['account_types']:
                context['account_types'][account_type] = {
                    'accounts': [],
                    'total': Decimal('0')
                }
            context['account_types'][account_type]['accounts'].append(account)
            context['account_types'][account_type]['total'] += account.current_balance

        return context


class AccountCreateView(LoginRequiredMixin, CreateView):
    """Create a new account."""
    model = Account
    form_class = AccountForm
    template_name = 'transactions/account_form.html'
    success_url = reverse_lazy('transactions:accounts')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Account created successfully!')
        return response


class AccountUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing account."""
    model = Account
    form_class = AccountForm
    template_name = 'transactions/account_form.html'
    success_url = reverse_lazy('transactions:accounts')

    def get_queryset(self):
        queryset = Account.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(owner=self.request.user, family_group__isnull=True)
        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Account updated successfully!')
        return response


class AccountDetailView(LoginRequiredMixin, DetailView):
    """Account detail view with recent transactions."""
    model = Account
    template_name = 'transactions/account_detail.html'
    context_object_name = 'account'

    def get_queryset(self):
        queryset = Account.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(owner=self.request.user, family_group__isnull=True)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get recent transactions for this account
        recent_transactions = Transaction.objects.filter(
            Q(account=self.object) | Q(to_account=self.object) | Q(from_account=self.object),
            is_active=True
        ).select_related('category', 'user').order_by('-date', '-created_at')[:20]

        context['recent_transactions'] = recent_transactions

        # Calculate monthly spending trends
        from django.db.models.functions import TruncMonth
        monthly_data = Transaction.objects.filter(
            account=self.object,
            transaction_type='expense',
            is_active=True,
            date__gte=timezone.now().date() - timedelta(days=365)
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            total=Sum('amount')
        ).order_by('month')

        context['monthly_spending'] = monthly_data

        return context


@login_required
def recurring_transactions_list(request):
    """List recurring transactions."""
    queryset = RecurringTransaction.objects.filter(is_active=True)

    if hasattr(request, 'current_family_group') and request.current_family_group:
        queryset = queryset.filter(family_group=request.current_family_group)
    else:
        queryset = queryset.filter(user=request.user, family_group__isnull=True)

    recurring_transactions = queryset.order_by('next_due_date')

    return render(request, 'transactions/recurring_list.html', {
        'recurring_transactions': recurring_transactions
    })


class RecurringTransactionCreateView(LoginRequiredMixin, CreateView):
    """Create a recurring transaction."""
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'transactions/recurring_form.html'
    success_url = reverse_lazy('transactions:recurring_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Recurring transaction created successfully!')
        return response


class RecurringTransactionUpdateView(LoginRequiredMixin, UpdateView):
    """Update a recurring transaction."""
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'transactions/recurring_form.html'
    success_url = reverse_lazy('transactions:recurring_list')

    def get_queryset(self):
        queryset = RecurringTransaction.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(user=self.request.user, family_group__isnull=True)
        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['family_group'] = getattr(self.request, 'current_family_group', None)
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Recurring transaction updated successfully!')
        return response


class RecurringTransactionDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a recurring transaction."""
    model = RecurringTransaction
    template_name = 'transactions/recurring_confirm_delete.html'
    success_url = reverse_lazy('transactions:recurring_list')

    def get_queryset(self):
        queryset = RecurringTransaction.objects.filter(is_active=True)
        if hasattr(self.request, 'current_family_group') and self.request.current_family_group:
            queryset = queryset.filter(family_group=self.request.current_family_group)
        else:
            queryset = queryset.filter(user=self.request.user, family_group__isnull=True)
        return queryset

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # Soft delete
        self.object.is_active = False
        self.object.save()
        messages.success(request, 'Recurring transaction deleted successfully!')
        return HttpResponseRedirect(self.success_url)


@login_required
def bulk_upload_transactions(request):
    """Bulk upload transactions from CSV/Excel files."""
    if request.method == 'POST':
        form = BulkTransactionUploadForm(
            request.POST,
            request.FILES,
            user=request.user,
            family_group=getattr(request, 'current_family_group', None)
        )

        if form.is_valid():
            file = form.cleaned_data['file']
            account = form.cleaned_data['account']
            has_header = form.cleaned_data['has_header']

            try:
                # Use the import service
                from .services import import_service

                result = import_service.import_transactions(
                    file=file,
                    account=account,
                    user=request.user,
                    family_group=getattr(request, 'current_family_group', None),
                    has_header=has_header
                )

                if result['success']:
                    if result['created_count'] > 0:
                        messages.success(
                            request,
                            f'Successfully imported {result["created_count"]} transactions.'
                        )

                    if result.get('errors'):
                        error_count = len(result['errors'])
                        error_message = f"Encountered {error_count} errors during import:"

                        # Show first 10 errors
                        for error in result['errors'][:10]:
                            error_message += f"\n• {error}"

                        if error_count > 10:
                            error_message += f"\n... and {error_count - 10} more errors."

                        messages.warning(request, error_message)

                    return redirect('transactions:list')
                else:
                    messages.error(request, f'Import failed: {result.get("error", "Unknown error")}')

            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')

    else:
        form = BulkTransactionUploadForm(
            user=request.user,
            family_group=getattr(request, 'current_family_group', None)
        )

    return render(request, 'transactions/bulk_upload.html', {'form': form})


@login_required
def download_template(request, format):
    """Download transaction import template in specified format."""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # Define template headers
    headers = [
        'Date (YYYY-MM-DD)',
        'Description',
        'Amount',
        'Type (income/expense)',
        'Category (optional)',
        'Notes (optional)'
    ]
    
    sample_data = [
        ['2025-01-01', 'Sample Income Transaction', '1000.00', 'income', 'Salary', 'Monthly salary payment'],
        ['2025-01-02', 'Sample Expense Transaction', '50.00', 'expense', 'Groceries', 'Weekly grocery shopping'],
        ['2025-01-03', 'Sample Transfer', '200.00', 'expense', 'Transfer', 'Money transfer to savings'],
    ]
    
    if format.lower() == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transaction_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_data)
        
        return response
        
    elif format.lower() == 'excel':
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transaction Template"
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Add sample data
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transaction_template.xlsx"'
        
        wb.save(response)
        return response
    
    else:
        messages.error(request, 'Invalid template format requested.')
        return redirect('transactions:bulk_upload')


@login_required
def bulk_delete_transactions(request):
    """Bulk delete selected transactions."""
    if request.method == 'POST':
        transaction_ids = request.POST.getlist('transaction_ids')
        
        if not transaction_ids:
            messages.warning(request, 'No transactions selected for deletion.')
            return redirect('transactions:list')
        
        try:
            # Get transactions belonging to the user
            queryset = Transaction.objects.filter(
                id__in=transaction_ids,
                user=request.user,
                is_active=True
            )
            
            # Apply family group filter if applicable
            if hasattr(request, 'current_family_group') and request.current_family_group:
                queryset = queryset.filter(family_group=request.current_family_group)
            else:
                queryset = queryset.filter(family_group__isnull=True)
            
            # Get affected accounts for balance updates
            affected_accounts = set()
            for transaction in queryset:
                affected_accounts.add(transaction.account)
                if transaction.to_account:
                    affected_accounts.add(transaction.to_account)
                if transaction.from_account:
                    affected_accounts.add(transaction.from_account)
            
            # Delete transactions
            deleted_count = queryset.count()
            queryset.delete()
            
            # Update account balances
            for account in affected_accounts:
                account.update_balance()
            
            messages.success(
                request, 
                f'Successfully deleted {deleted_count} transaction{"s" if deleted_count != 1 else ""}.'
            )
            
        except Exception as e:
            logger.error(f"Error in bulk delete: {str(e)}")
            messages.error(request, 'An error occurred while deleting transactions. Please try again.')
    
    return redirect('transactions:list')