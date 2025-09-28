"""
Services for transaction import/export and bulk operations.
"""
import csv
import io
import openpyxl
import logging
import re
import PyPDF2
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Tuple, Optional
from django.db import transaction
from django.contrib.auth import get_user_model
from .models import Account, Transaction
from ..core.models import Category

# Import modular bank analyzers
try:
    import sys
    import os
    # Add the project root to Python path
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    sys.path.insert(0, project_root)
    from bank_pdf_analyzers import BankAnalyzerFactory
    MODULAR_ANALYZERS_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Modular bank analyzers not available: {e}")
    MODULAR_ANALYZERS_AVAILABLE = False

logger = logging.getLogger(__name__)
User = get_user_model()


class TransactionImportService:
    """Service for importing transactions from various file formats."""

    def __init__(self):
        self.supported_formats = ['.csv', '.xlsx', '.xls', '.pdf']
        self.required_columns = ['date', 'description', 'amount', 'type']
        self.date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']

        # PDF parsing patterns
        self.pdf_date_patterns = [
            r'\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b',  # MM/DD/YYYY or DD/MM/YYYY
            r'\b(\d{4}-\d{1,2}-\d{1,2})\b',           # YYYY-MM-DD
            r'\b(\d{1,2}\s+\w{3}\s+\d{4})\b',         # DD MMM YYYY
            r'\b(\w{3}\s+\d{1,2},?\s+\d{4})\b'        # MMM DD, YYYY
        ]

        self.pdf_amount_patterns = [
            r'[-+]?\$?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',  # Currency with commas
            r'[-+]?\s*(\d+\.\d{2})',                          # Decimal amounts
            r'\((\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\)',          # Parentheses for negative
        ]

    def import_transactions(
        self,
        file,
        account: Account,
        user,  # User model instance
        family_group=None,
        has_header: bool = True
    ) -> Dict[str, any]:
        """
        Import transactions from a file.

        Args:
            file: Uploaded file object
            account: Account to associate transactions with
            user: User importing the transactions
            family_group: Optional family group context
            has_header: Whether the file has a header row

        Returns:
            Dict with success status, created count, and errors
        """
        try:
            file_name = file.name.lower()

            if not any(file_name.endswith(fmt) for fmt in self.supported_formats):
                return {
                    'success': False,
                    'error': f'Unsupported file format. Supported formats: {", ".join(self.supported_formats)}',
                    'created_count': 0,
                    'errors': []
                }

            # Parse file based on format
            if file_name.endswith('.csv'):
                return self._import_csv(file, account, user, family_group, has_header)
            elif file_name.endswith(('.xlsx', '.xls')):
                return self._import_excel(file, account, user, family_group, has_header)
            elif file_name.endswith('.pdf'):
                return self._import_pdf(file, account, user, family_group, has_header)
            else:
                return {
                    'success': False,
                    'error': 'Unsupported file format',
                    'created_count': 0,
                    'errors': []
                }

        except Exception as e:
            logger.error(f"Error importing transactions: {str(e)}")
            return {
                'success': False,
                'error': f'Import failed: {str(e)}',
                'created_count': 0,
                'errors': []
            }

    def _import_csv(
        self,
        file,
        account: Account,
        user,  # User model instance
        family_group=None,
        has_header: bool = True
    ) -> Dict[str, any]:
        """Import transactions from CSV file."""
        transactions_created = 0
        errors = []

        try:
            # Read and decode file
            file.seek(0)
            decoded_file = file.read().decode('utf-8-sig')  # Handle BOM
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)

            # Skip header if present
            if has_header:
                headers = next(reader, None)
                if not headers:
                    return {
                        'success': False,
                        'error': 'File is empty or invalid',
                        'created_count': 0,
                        'errors': []
                    }

            # Process rows in batches
            batch_size = 100
            transactions_to_create = []
            row_num = 2 if has_header else 1

            for row in reader:
                try:
                    parsed_transaction = self._parse_csv_row(
                        row, row_num, account, user, family_group
                    )
                    if parsed_transaction:
                        transactions_to_create.append(parsed_transaction)

                    # Create batch when size reached
                    if len(transactions_to_create) >= batch_size:
                        created = self._create_transaction_batch(transactions_to_create)
                        transactions_created += created
                        transactions_to_create = []

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

                row_num += 1

            # Create remaining transactions
            if transactions_to_create:
                created = self._create_transaction_batch(transactions_to_create)
                transactions_created += created

            # Update account balance
            account.update_balance()

            return {
                'success': True,
                'created_count': transactions_created,
                'errors': errors
            }

        except Exception as e:
            logger.error(f"CSV import error: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'created_count': transactions_created,
                'errors': errors
            }

    def _import_excel(
        self,
        file,
        account: Account,
        user,  # User model instance
        family_group=None,
        has_header: bool = True
    ) -> Dict[str, any]:
        """Import transactions from Excel file."""
        transactions_created = 0
        errors = []

        try:
            # Load workbook
            file.seek(0)
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active

            # Skip header if present
            start_row = 2 if has_header else 1

            # Process rows in batches
            batch_size = 100
            transactions_to_create = []
            row_num = start_row

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                try:
                    if not row or all(cell is None for cell in row):
                        continue

                    parsed_transaction = self._parse_excel_row(
                        row, row_num, account, user, family_group
                    )
                    if parsed_transaction:
                        transactions_to_create.append(parsed_transaction)

                    # Create batch when size reached
                    if len(transactions_to_create) >= batch_size:
                        created = self._create_transaction_batch(transactions_to_create)
                        transactions_created += created
                        transactions_to_create = []

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

                row_num += 1

            # Create remaining transactions
            if transactions_to_create:
                created = self._create_transaction_batch(transactions_to_create)
                transactions_created += created

            wb.close()

            # Update account balance
            account.update_balance()

            return {
                'success': True,
                'created_count': transactions_created,
                'errors': errors
            }

        except Exception as e:
            logger.error(f"Excel import error: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'created_count': transactions_created,
                'errors': errors
            }

    def _parse_csv_row(
        self,
        row: List[str],
        row_num: int,
        account: Account,
        user,  # User model instance
        family_group=None
    ) -> Optional[Dict]:
        """Parse a single CSV row into transaction data."""
        if len(row) < 4:
            raise ValueError("Insufficient columns (need at least: date, description, amount, type)")

        # Parse required fields
        date_str = row[0].strip() if row[0] else ''
        description = row[1].strip() if row[1] else ''
        amount_str = row[2].strip() if row[2] else ''
        trans_type = row[3].strip().lower() if row[3] else ''

        if not all([date_str, description, amount_str, trans_type]):
            raise ValueError("Missing required data")

        # Parse and validate date
        transaction_date = self._parse_date(date_str)
        if not transaction_date:
            raise ValueError(f"Invalid date format: {date_str}")

        # Parse and validate amount
        try:
            amount = Decimal(amount_str.replace('$', '').replace('₹', '').replace(',', '').replace('(', '-').replace(')', ''))
            if amount <= 0:
                raise ValueError("Amount must be positive")
        except (InvalidOperation, ValueError):
            raise ValueError(f"Invalid amount format: {amount_str}")

        # Validate transaction type
        if trans_type not in ['income', 'expense']:
            raise ValueError(f"Invalid transaction type: {trans_type}")

        # Parse optional category
        category = None
        if len(row) > 4 and row[4] and row[4].strip():
            category_name = row[4].strip()
            try:
                # Try to find existing category first
                if family_group:
                    category = Category.objects.filter(
                        name__iexact=category_name,
                        family_group=family_group,
                        category_type=trans_type
                    ).first()
                else:
                    category = Category.objects.filter(
                        name__iexact=category_name,
                        family_group__isnull=True,
                        category_type=trans_type,
                        is_system_category=False
                    ).first()

                # Create category if not found
                if not category:
                    category = Category.objects.create(
                        name=category_name,
                        category_type=trans_type,
                        family_group=family_group,
                        color='#007bff'
                    )
            except Exception as e:
                logger.warning(f"Could not create category '{category_name}': {str(e)}")
                category = None

        return {
            'amount': amount,
            'description': description,
            'transaction_type': trans_type,
            'category': category,
            'account': account,
            'date': transaction_date,
            'user': user,
            'family_group': family_group,
            'imported_from': 'csv_upload'
        }

    def _parse_excel_row(
        self,
        row: tuple,
        row_num: int,
        account: Account,
        user,  # User model instance
        family_group=None
    ) -> Optional[Dict]:
        """Parse a single Excel row into transaction data."""
        if len(row) < 4:
            return None

        # Parse required fields
        date_cell = row[0]
        description = str(row[1]).strip() if row[1] else ''
        amount_str = str(row[2]).strip() if row[2] else ''
        trans_type = str(row[3]).strip().lower() if row[3] else ''

        if not all([date_cell, description, amount_str, trans_type]):
            raise ValueError("Missing required data")

        # Parse date (handle both datetime and string)
        if isinstance(date_cell, datetime):
            transaction_date = date_cell.date()
        else:
            transaction_date = self._parse_date(str(date_cell))
            if not transaction_date:
                raise ValueError(f"Invalid date format: {date_cell}")

        # Parse and validate amount
        try:
            amount = Decimal(str(amount_str).replace('$', '').replace('₹', '').replace(',', '').replace('(', '-').replace(')', ''))
            if amount <= 0:
                raise ValueError("Amount must be positive")
        except (InvalidOperation, ValueError):
            raise ValueError(f"Invalid amount format: {amount_str}")

        # Validate transaction type
        if trans_type not in ['income', 'expense']:
            raise ValueError(f"Invalid transaction type: {trans_type}")

        # Parse optional category (for Excel)
        category = None
        if len(row) > 4 and row[4]:
            category_name = str(row[4]).strip()
            if category_name:
                try:
                    # Try to find existing category first
                    if family_group:
                        category = Category.objects.filter(
                            name__iexact=category_name,
                            family_group=family_group,
                            category_type=trans_type
                        ).first()
                    else:
                        category = Category.objects.filter(
                            name__iexact=category_name,
                            family_group__isnull=True,
                            category_type=trans_type,
                            is_system_category=False
                        ).first()

                    # Create category if not found
                    if not category:
                        category = Category.objects.create(
                            name=category_name,
                            category_type=trans_type,
                            family_group=family_group,
                            color='#007bff'
                        )
                except Exception as e:
                    logger.warning(f"Could not create category '{category_name}': {str(e)}")
                    category = None

        return {
            'amount': amount,
            'description': description,
            'transaction_type': trans_type,
            'category': category,
            'account': account,
            'date': transaction_date,
            'user': user,
            'family_group': family_group,
            'imported_from': 'excel_upload'
        }

    def _parse_date(self, date_str: str):
        """Parse date string using multiple formats with bank-specific handling."""
        if not date_str:
            logger.error("Empty date string provided to _parse_date")
            return None
            
        date_str = date_str.strip()
        logger.debug(f"Parsing date string: '{date_str}'")
        
        # Enhanced format list with bank-specific priorities
        # YYYY-MM-DD first for converted dates, then DD/MM/YYYY for original formats
        priority_formats = [
            '%Y-%m-%d',        # For converted dates (HDFC, SBI, Axis)
            '%d/%m/%Y',        # DD/MM/YYYY format (Federal Bank, original HDFC)
            '%d/%m/%y',        # DD/MM/YY format (HDFC 2-digit year)
            '%Y-%m-%d %H:%M:%S',  # With timestamp
            '%m/%d/%Y',        # MM/DD/YYYY format (US style)
            '%d-%m-%Y',        # DD-MM-YYYY format
            '%d %b %Y',        # DD MMM YYYY format (SBI)
            '%d-%b-%Y'         # DD-MMM-YYYY format
        ]
        
        for date_format in priority_formats:
            try:
                parsed_date = datetime.strptime(date_str, date_format).date()
                logger.debug(f"Successfully parsed '{date_str}' using format '{date_format}' -> {parsed_date}")
                return parsed_date
            except ValueError:
                continue
                
        # If no format worked, log the issue with more details
        logger.error(f"Failed to parse date: '{date_str}' using any known format")
        logger.error(f"Attempted formats: {priority_formats}")
        return None

    def _create_transaction_batch(self, transactions: List[Dict]) -> int:
        """Create transactions in batch for better performance with duplicate detection."""
        created_count = 0
        skipped_duplicates = 0
        
        try:
            # Create transactions one by one to trigger signals and validation
            with transaction.atomic():
                for trans_data in transactions:
                    try:
                        # Check for duplicates based on date, description (first 50 chars), amount, and account
                        description_part = trans_data.get('description', '')[:50]  # First 50 chars
                        
                        existing_transaction = Transaction.objects.filter(
                            date=trans_data.get('date'),
                            description__icontains=description_part,
                            amount=trans_data.get('amount'),
                            account=trans_data.get('account')
                        ).first()
                        
                        if existing_transaction:
                            logger.info(f"Skipping duplicate transaction: {description_part} - {trans_data.get('amount')} on {trans_data.get('date')}")
                            skipped_duplicates += 1
                            continue
                            
                        Transaction.objects.create(**trans_data)
                        created_count += 1
                        
                    except Exception as e:
                        logger.error(f"Error creating individual transaction: {str(e)}")
                        # Continue with other transactions
                        continue

            if skipped_duplicates > 0:
                logger.info(f"Skipped {skipped_duplicates} duplicate transactions during bulk import")
                
            return created_count
            
        except Exception as e:
            logger.error(f"Error in transaction batch creation: {str(e)}")
            return created_count

    def _import_pdf(
        self,
        file,
        account: Account,
        user,  # User model instance
        family_group=None,
        has_header: bool = True
    ) -> Dict[str, any]:
        """Import transactions from PDF file (bank statements)."""
        transactions_created = 0
        errors = []

        try:
            # Read PDF content
            file.seek(0)
            
            # Try to create PDF reader with error handling
            try:
                pdf_reader = PyPDF2.PdfReader(file)
            except Exception as pdf_error:
                logger.error(f"Failed to read PDF file: {pdf_error}")
                return {
                    'success': False,
                    'error': (
                        f'❌ Cannot read PDF file\n\n'
                        f'🔍 Error details: {str(pdf_error)}\n\n'
                        '💡 This usually means:\n'
                        '• PDF file is corrupted or damaged\n'
                        '• PDF file is not a valid PDF format\n'
                        '• PDF was created incorrectly\n\n'
                        '🔄 Solutions:\n'
                        '1. Re-download the statement from HDFC NetBanking\n'
                        '2. Ensure you\'re downloading the PDF (not printing to PDF)\n'
                        '3. Try a different date range or account\n'
                        '4. Use CSV format as an alternative\n\n'
                        '📞 If the problem persists, contact HDFC support'
                    ),
                    'created_count': 0,
                    'errors': []
                }

            # Check if PDF is encrypted
            if pdf_reader.is_encrypted:
                return {
                    'success': False,
                    'error': 'PDF is password protected. Please provide an unprotected PDF.',
                    'created_count': 0,
                    'errors': []
                }

            # Extract text from all pages with enhanced extraction
            pdf_text = ""
            total_pages = len(pdf_reader.pages)
            logger.info(f"Processing PDF with {total_pages} pages")
            
            extraction_attempts = []
            
            for page_num in range(total_pages):
                page_text = ""
                attempt_info = f"Page {page_num + 1}"
                
                try:
                    page = pdf_reader.pages[page_num]
                    logger.info(f"Extracting text from page {page_num + 1}")
                    
                    # Method 1: Standard text extraction
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            attempt_info += f" - Standard extraction: {len(page_text)} chars"
                        else:
                            attempt_info += " - Standard extraction: No text"
                    except Exception as e:
                        attempt_info += f" - Standard extraction failed: {e}"
                    
                    # Method 2: Alternative extraction methods
                    if not page_text or len(page_text.strip()) < 20:
                        logger.info(f"Trying alternative extraction methods for page {page_num + 1}")
                        
                        # Try extractText (older PyPDF2 method)
                        try:
                            if hasattr(page, 'extractText'):
                                alt_text = page.extractText()
                                if alt_text and len(alt_text) > len(page_text):
                                    page_text = alt_text
                                    attempt_info += f" - Alt method 1: {len(alt_text)} chars"
                        except Exception as e:
                            attempt_info += f" - Alt method 1 failed: {e}"
                        
                        # Try accessing text content directly
                        try:
                            if hasattr(page, '_get_contents_of_obj'):
                                # This is a more advanced method
                                pass
                        except:
                            pass
                    
                    # Method 3: Check if page has any content objects
                    try:
                        if hasattr(page, '/Contents'):
                            contents = page['/Contents']
                            if contents:
                                attempt_info += " - Has content objects"
                            else:
                                attempt_info += " - No content objects"
                    except:
                        pass
                    
                    if page_text and page_text.strip():
                        pdf_text += f"\n--- PAGE {page_num + 1} ---\n" + page_text + "\n"
                        logger.info(f"Successfully extracted {len(page_text)} characters from page {page_num + 1}")
                    else:
                        logger.warning(f"No meaningful text extracted from page {page_num + 1}")
                        
                except Exception as e:
                    logger.error(f"Critical error extracting from page {page_num + 1}: {str(e)}")
                    attempt_info += f" - Critical error: {e}"
                
                extraction_attempts.append(attempt_info)
            
            # Log extraction summary
            logger.info("PDF Extraction Summary:")
            for attempt in extraction_attempts:
                logger.info(f"  {attempt}")

            # Validate extracted text
            pdf_text = pdf_text.strip()
            if not pdf_text:
                logger.error("No text could be extracted from any page")
                logger.error("PDF Extraction Details:")
                for attempt in extraction_attempts:
                    logger.error(f"  {attempt}")
                
                return {
                    'success': False,
                    'error': (
                        '❌ No text could be extracted from PDF\n\n'
                        '🔍 This usually happens with:\n'
                        '• Image-based or scanned PDFs (most common)\n'
                        '• Corrupted or damaged PDF files\n'
                        '• Password-protected PDFs\n'
                        '• PDFs created from images/screenshots\n\n'
                        '💡 Solutions for HDFC Bank Statements:\n'
                        '1. Log into HDFC NetBanking\n'
                        '2. Go to "Account Summary" → "View Statement"\n'
                        '3. Select your account and date range\n'
                        '4. Click "Download Statement" (not "Print")\n'
                        '5. Choose "PDF Format" (not "Image Format")\n'
                        '6. Ensure the downloaded PDF has selectable text\n\n'
                        '🔄 Alternative options:\n'
                        '• Try downloading from HDFC Mobile Banking app\n'
                        '• Request statement via email from bank\n'
                        '• Use CSV format if PDF continues to fail\n\n'
                        '⚠️  Note: Screenshot PDFs and scanned images won\'t work'
                    ),
                    'created_count': 0,
                    'errors': []
                }
                
            # Check for meaningful content
            meaningful_chars = sum(1 for c in pdf_text if c.isalnum())
            text_lines = [line.strip() for line in pdf_text.split('\n') if line.strip()]
            
            if meaningful_chars < 10:
                logger.error(f"PDF contains insufficient readable text ({meaningful_chars} characters)")
                return {
                    'success': False,
                    'error': (
                        f'PDF contains insufficient readable text ({meaningful_chars} characters). '
                        'This indicates the PDF is likely a scanned image.\n\n'
                        'For HDFC bank statements:\n'
                        '1. Log into HDFC NetBanking\n'
                        '2. Go to Account Statements\n'
                        '3. Select "Download as PDF (Text Format)"\n'
                        '4. Avoid "Print" or "Image" format options'
                    ),
                    'created_count': 0,
                    'errors': []
                }
            
            # Additional validation for bank statement format
            bank_indicators = ['statement', 'account', 'balance', 'transaction', 'debit', 'credit']
            has_bank_content = any(indicator.lower() in pdf_text.lower() for indicator in bank_indicators)
            
            if not has_bank_content and len(text_lines) < 5:
                logger.warning(f"PDF doesn't appear to contain bank statement data")
                return {
                    'success': False,
                    'error': (
                        'This PDF doesn\'t appear to contain bank statement data.\n\n'
                        'Please ensure you\'re uploading:\n'
                        '• A bank account statement (not a passbook scan)\n'
                        '• A complete statement with transaction details\n'
                        '• A text-based PDF (not a scanned image)'
                    ),
                    'created_count': 0,
                    'errors': []
                }
                
            logger.info(f"Successfully extracted {len(pdf_text)} characters, {meaningful_chars} meaningful characters")

            # Debug: Log extracted text (first 1000 characters)
            logger.info(f"PDF text extracted ({len(pdf_text)} chars): {pdf_text[:1000]}...")

            # Extract statement date if available
            statement_date = self._extract_statement_date(pdf_text)

            # Parse transactions from text
            transactions_data = self._parse_pdf_transactions(pdf_text, statement_date)

            logger.info(f"Found {len(transactions_data)} potential transactions in PDF")

            if not transactions_data:
                # Provide more detailed feedback
                lines_with_numbers = [line for line in pdf_text.split('\n')
                                    if re.search(r'\d+\.\d{2}', line) and len(line.strip()) > 10]

                debug_info = f"Lines with amounts found: {len(lines_with_numbers)}"
                if lines_with_numbers:
                    debug_info += f"\nSample lines:\n" + "\n".join(lines_with_numbers[:5])

                return {
                    'success': False,
                    'error': f'No transaction data could be found in PDF. {debug_info}',
                    'created_count': 0,
                    'errors': []
                }

            # Process parsed transactions
            batch_size = 100
            transactions_to_create = []
            row_num = 1

            for trans_data in transactions_data:
                try:
                    parsed_transaction = self._create_pdf_transaction(
                        trans_data, row_num, account, user, family_group
                    )
                    if parsed_transaction:
                        transactions_to_create.append(parsed_transaction)

                    # Create batch when size reached
                    if len(transactions_to_create) >= batch_size:
                        created = self._create_transaction_batch(transactions_to_create)
                        transactions_created += created
                        transactions_to_create = []

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

                row_num += 1

            # Create remaining transactions
            if transactions_to_create:
                created = self._create_transaction_batch(transactions_to_create)
                transactions_created += created

            # Update account balance
            account.update_balance()

            return {
                'success': True,
                'created_count': transactions_created,
                'errors': errors
            }

        except Exception as e:
            logger.error(f"PDF import error: {str(e)}")
            return {
                'success': False,
                'error': f'PDF processing failed: {str(e)}',
                'created_count': transactions_created,
                'errors': errors
            }

    def _extract_statement_date(self, pdf_text: str) -> str:
        """Extract statement date from PDF text with improved logic."""
        logger.info("=== EXTRACTING STATEMENT DATE ===")
        
        # Enhanced date patterns for bank statements
        date_patterns = [
            r'Date of Issue\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
            r'Statement Date\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
            r'Generated on\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
            r'Date\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
            r'Transaction Date\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
            r'(\d{2}/\d{2}/\d{4})\s+to\s+(\d{2}/\d{2}/\d{4})',  # Date range
            r'As on\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})',
        ]

        # Look for date patterns
        for i, pattern in enumerate(date_patterns, 1):
            matches = re.findall(pattern, pdf_text, re.IGNORECASE)
            if matches:
                if isinstance(matches[0], tuple):  # Date range pattern
                    found_date = matches[0][1]  # Use end date
                else:
                    found_date = matches[0]
                logger.info(f"Found statement date using pattern {i}: '{found_date}'")
                return found_date

        # Extract all dates and use the most recent looking one
        all_dates = re.findall(r'\b(\d{1,2}/\d{1,2}/\d{4})\b', pdf_text)
        if all_dates:
            # Filter dates that look like transaction dates (not too old)
            from datetime import datetime, timedelta
            current_year = datetime.now().year
            valid_dates = []
            
            for date_str in all_dates:
                try:
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    # Only consider dates from last 2 years
                    if date_obj.year >= current_year - 1:
                        valid_dates.append((date_str, date_obj))
                except ValueError:
                    continue
            
            if valid_dates:
                # Sort by date and take the most recent
                valid_dates.sort(key=lambda x: x[1], reverse=True)
                found_date = valid_dates[0][0]
                logger.info(f"Using most recent valid date: '{found_date}'")
                return found_date

        # Final fallback - use current date
        fallback_date = datetime.now().strftime('%d/%m/%Y')
        logger.warning(f"No valid statement date found, using current date: {fallback_date}")
        return fallback_date

    def _parse_pdf_transactions(self, pdf_text: str, statement_date: str = None) -> List[Dict]:
        """Extract transaction data from PDF text - NEW MODULAR BANK ANALYZER SYSTEM."""
        logger.info(f"=== MODULAR BANK ANALYZER PARSING STARTED ===")
        logger.info(f"Statement date: {statement_date}")
        logger.info(f"Processing {len(pdf_text)} characters from PDF")

        # Try new modular analyzer system first
        if MODULAR_ANALYZERS_AVAILABLE:
            try:
                analyzer = BankAnalyzerFactory.get_analyzer(pdf_text)
                if analyzer:
                    logger.info(f"Using modular analyzer: {analyzer.bank_name}")
                    transactions = analyzer.parse_transactions(pdf_text)
                    logger.info(f"Modular analyzer found {len(transactions)} transactions")
                    
                    # Convert to expected format and add classifications
                    processed_transactions = []
                    for txn in transactions:
                        # Ensure proper classification
                        final_type = analyzer.classify_transaction(
                            txn['description'], 
                            txn['amount'], 
                            txn
                        )
                        
                        processed_transactions.append({
                            'date': txn['date'],
                            'description': txn['description'],
                            'amount': abs(txn['amount']),  # Always positive
                            'type': final_type,
                            'balance': txn.get('balance'),
                            'bank': txn.get('bank', 'Unknown'),
                            'raw_line': txn.get('raw_line', '')
                        })
                    
                    return processed_transactions
                else:
                    logger.warning("No modular analyzer found, falling back to legacy system")
            except Exception as e:
                logger.error(f"Modular analyzer failed: {e}")
                logger.info("Falling back to legacy parsing system")

        # Fallback to legacy system
        transactions = []
        lines = pdf_text.split('\n')
        logger.info(f"Using legacy parser for {len(lines)} lines")

        # Detect bank type from PDF content
        bank_type = self._detect_bank_type(pdf_text)
        logger.info(f"Legacy detected bank type: {bank_type}")

        if bank_type == 'FEDERAL':
            return self._parse_federal_bank_transactions(lines, statement_date)
        elif bank_type == 'SBI':
            return self._parse_sbi_transactions(lines, statement_date)
        elif bank_type == 'HDFC':
            return self._parse_hdfc_transactions(lines, statement_date)
        elif bank_type == 'AXIS':
            return self._parse_axis_transactions(lines, statement_date)
        else:
            logger.warning(f"Unknown bank type, trying generic parsing")
            return self._parse_generic_bank_transactions(lines, statement_date)

    def _detect_bank_type(self, pdf_text: str) -> str:
        """Detect bank type from PDF content with improved accuracy."""
        text_lower = pdf_text.lower()
        
        # Score-based detection for better accuracy
        bank_scores = {
            'FEDERAL': 0,
            'SBI': 0,
            'HDFC': 0,
            'AXIS': 0
        }
        
        # Federal Bank indicators (strong indicators get higher scores)
        federal_indicators = [
            ('federal bank limited', 5),
            ('federal bank', 3),
            ('federal towers', 4),
            ('fdrl', 3),
            ('fdrlinbb', 4)
        ]
        
        for indicator, score in federal_indicators:
            if indicator in text_lower:
                bank_scores['FEDERAL'] += score
        
        # SBI indicators
        sbi_indicators = [
            ('state bank of india', 5),
            ('sbin0020312', 4),  # Specific IFSC from the PDF
            ('state bank', 3),
            ('sbi', 2),
            ('sbin0', 3)
        ]
        
        for indicator, score in sbi_indicators:
            if indicator in text_lower:
                bank_scores['SBI'] += score
        
        # HDFC indicators
        hdfc_indicators = [
            ('hdfc bank limited', 5),
            ('housing development finance corporation', 5),
            ('hdfc bank', 4),
            ('hdfc0', 3)
        ]
        
        for indicator, score in hdfc_indicators:
            if indicator in text_lower:
                bank_scores['HDFC'] += score
        
        # Axis Bank indicators
        axis_indicators = [
            ('axis bank limited', 5),
            ('axis account no', 4),
            ('statement of axis account', 5),
            ('axis bank', 4),
            ('utib0004080', 4),  # Specific IFSC from the PDF
            ('utib', 3)
        ]
        
        for indicator, score in axis_indicators:
            if indicator in text_lower:
                bank_scores['AXIS'] += score
        
        # Additional context-based scoring
        # Look for account statements patterns
        if 'statement of axis account' in text_lower:
            bank_scores['AXIS'] += 10
        elif 'state bank of india' in text_lower and 'account number' in text_lower:
            bank_scores['SBI'] += 8
        elif 'hdfc bank' in text_lower and 'statement' in text_lower:
            bank_scores['HDFC'] += 8
        elif 'federal bank' in text_lower and 'statement' in text_lower:
            bank_scores['FEDERAL'] += 8
        
        # Return the bank with highest score (minimum threshold of 3)
        max_score = max(bank_scores.values())
        if max_score >= 3:
            for bank, score in bank_scores.items():
                if score == max_score:
                    return bank
        
        return 'GENERIC'

    def _parse_federal_bank_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse Federal Bank specific format."""
        transactions = []
        
        logger.info(f"=== FEDERAL BANK PARSING ===")
        
        # Federal Bank specific patterns
        federal_bank_patterns = [
            r'^(.+?TFR\s+[A-Z0-9]+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(Cr|Dr)\s*(.*)$'
        ]

        # Track parsing state
        current_date = None
        current_description = ""
        previous_balance = 0.0

        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if len(line) < 10:
                continue

            # Skip Federal Bank header/footer lines
            skip_indicators = [
                'federal bank', 'corporate office', 'statement of account',
                'opening balance', 'grand total', 'abbreviations', 'disclaimer',
                'page ', 'name :', 'communication address', 'ifsc', 'micr', 'swift'
            ]
            
            if any(skip in line.lower() for skip in skip_indicators):
                continue

            # Federal Bank date format: "22-MAY-2023 22-MAY-2023 IFN/..."
            date_desc_match = re.search(r'^(\d{1,2}-[A-Z]{3}-\d{4})\s+\d{1,2}-[A-Z]{3}-\d{4}\s+(.+)$', line)
            if date_desc_match:
                current_date = self._convert_federal_date(date_desc_match.group(1))
                current_description = date_desc_match.group(2).strip()
                continue

            # Check for Federal Bank transaction pattern
            for pattern_num, pattern in enumerate(federal_bank_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match and current_date:
                    transaction_data = self._process_federal_bank_transaction(
                        match, current_date, current_description, previous_balance, line_num
                    )
                    
                    if transaction_data:
                        transactions.append(transaction_data)
                        previous_balance = transaction_data.get('new_balance', previous_balance)
                        current_description = ""
                        current_date = None
                    break

        logger.info(f"=== FOUND {len(transactions)} FEDERAL BANK TRANSACTIONS ===")
        return transactions

    def _parse_sbi_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse SBI (State Bank of India) specific format - Improved for actual PDF format."""
        transactions = []
        
        logger.info(f"=== SBI PARSING (IMPROVED) ===")
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if len(line) < 20:
                continue

            # Skip SBI header/footer lines
            skip_indicators = [
                'state bank of india', 'account name', 'account number', 'branch',
                'ifsc code', 'micr code', 'customer id', 'nominee registered',
                'date credit balance details', 'ref no./cheque no', 'debit',
                'drawing power', 'interest rate', 'address', 'cif no', 'ckyc no'
            ]
            
            if any(skip in line.lower() for skip in skip_indicators):
                continue

            logger.info(f"SBI Line {line_num}: {line}")

            # SBI actual format from PDF: "100.00 - 01 JUN 2024 TRANSFER TO 4897695162091 -UPI/DR/415388277978/MOHAMMAD/KKBK/mohammadmu/UPI9.13"
            # Pattern: AMOUNT - DATE DESCRIPTION BALANCE
            sbi_patterns = [
                # Pattern 1: Amount - Date Description Balance
                r'^(\d+\.?\d*)\s*-\s*(\d{2}\s+[A-Z]{3}\s+\d{4})\s+(.+?)\s+([\d,]+\.?\d*)\s*$',
                # Pattern 2: - Amount Date Description Balance  
                r'^-\s+(\d+\.?\d*)\s+(\d{2}\s+[A-Z]{3}\s+\d{4})\s+(.+?)\s+([\d,]+\.?\d*)\s*$',
                # Pattern 3: Date Description Amount Balance
                r'^(\d{2}\s+[A-Z]{3}\s+\d{4})\s+(.+?)\s+(\d+\.?\d*)\s+([\d,]+\.?\d*)\s*$'
            ]
            
            transaction_found = False
            
            for pattern_num, pattern in enumerate(sbi_patterns, 1):
                match = re.search(pattern, line)
                if match:
                    try:
                        if pattern_num == 1:  # Amount - Date Description Balance
                            transaction_amount = match.group(1)
                            date_str = self._convert_sbi_date_format2(match.group(2))
                            description = match.group(3).strip()
                            balance = match.group(4)
                            
                            # Determine transaction type from description
                            desc_lower = description.lower()
                            if 'transfer to' in desc_lower or 'upi/dr' in desc_lower:
                                trans_type = 'expense'
                            else:
                                trans_type = 'income'
                                
                        elif pattern_num == 2:  # - Amount Date Description Balance
                            transaction_amount = match.group(1)
                            date_str = self._convert_sbi_date_format2(match.group(2))
                            description = match.group(3).strip()
                            balance = match.group(4)
                            
                            # Credit transaction (income)
                            trans_type = 'income'
                            
                        else:  # Pattern 3: Date Description Amount Balance
                            date_str = self._convert_sbi_date_format2(match.group(1))
                            description = match.group(2).strip()
                            transaction_amount = match.group(3)
                            balance = match.group(4)
                            
                            # Determine type from description
                            desc_lower = description.lower()
                            if any(term in desc_lower for term in ['transfer to', 'upi/dr', 'withdrawal', 'debit']):
                                trans_type = 'expense'
                            else:
                                trans_type = 'income'
                        
                        # Clean amount
                        transaction_amount = transaction_amount.replace(',', '')
                        
                        if float(transaction_amount) > 0:
                            logger.info(f"*** SBI TRANSACTION FOUND ON LINE {line_num} ***")
                            logger.info(f"  Date: {date_str}")
                            logger.info(f"  Description: {description}")
                            logger.info(f"  Amount: {transaction_amount}")
                            logger.info(f"  Type: {trans_type}")
                            
                            transactions.append({
                                'date_str': date_str,
                                'description': description,
                                'amount_str': transaction_amount,
                                'type': trans_type,
                                'source_line': line,
                                'pattern_used': pattern_num,
                                'bank_type': 'SBI'
                            })
                            
                            transaction_found = True
                            break
                            
                    except Exception as e:
                        logger.error(f"Error parsing SBI transaction on line {line_num}: {str(e)}")
                        continue

        logger.info(f"=== FOUND {len(transactions)} SBI TRANSACTIONS ===")
        return transactions

    def _parse_hdfc_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse HDFC Bank specific format with enhanced multi-line and comprehensive pattern matching."""
        transactions = []
        
        logger.info(f"=== HDFC PARSING ===")
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            line_num = i + 1
            
            if len(line) < 8:
                i += 1
                continue

            # Skip HDFC header/footer lines - made more specific to avoid false positives
            skip_indicators = [
                'hdfc bank', 'housing development finance', 'statement of account',
                'account number:', 'branch:', 'customer name:', 'ifsc code:',
                'opening balance', 'closing balance', 'generated on:',
                'cheque no', 'ref no', 'value dt', 'withdrawal amt', 'deposit amt',
                'this is a computer generated', 'contents of this statement',
                'mr ', 'joint holders', 'nomination', 'address', 'city', 'state',
                'phone no', 'email', 'cust id', 'account status', 'rtgs/neft ifsc'
            ]
            
            # Additional specific checks to avoid filtering valid transactions
            # Skip lines that are ONLY page numbers or headers
            if (line.lower().startswith('page no') or 
                line.lower() == 'page no' or
                re.match(r'^\s*page\s+no\s*[:.]?\s*\d+\s*$', line, re.IGNORECASE)):
                i += 1
                continue
            
            if any(skip in line.lower() for skip in skip_indicators):
                i += 1
                continue

            logger.info(f"HDFC Line {line_num}: {line}")

            # Enhanced HDFC patterns for comprehensive transaction capture
            hdfc_patterns = [
                # Pattern 1: Full HDFC format - DD/MM/YY Description RefNo DD/MM/YY Amount Balance
                r'^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(\d{10,})\s+\d{2}/\d{2}/\d{2}\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 2: DD/MM/YYYY format with 4-digit year 
                r'^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 3: DD/MM/YY simple format with two amounts (transaction amount and balance)
                r'^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 4: With reference number between description and amounts
                r'^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(\d{8,})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 5: HDFC format with description, ref number, negative amount, balance 
                r'^(\d{2}/\d{2}/\d{2,4})\s+(.+?)\s+(\d{8,})\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 6: UPI/ATM/POS specific pattern with negative amounts
                r'^(\d{2}/\d{2}/\d{2,4})\s+(UPI-|ATM|ATW|POS).+?\s+(\d{8,})\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 7: Date + Description + single negative amount + balance (common HDFC format)
                r'^(\d{2}/\d{2}/\d{2,4})\s+(.+?)\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})$',
                # Pattern 8: ATM/EAW transactions with specific format (captures missing June 27th transaction)
                r'^(\d{2}/\d{2}/\d{2,4})\s+(EAW-|ATW-|ATM-)(.+?)\s+(\d{8,})\s+\d{2}/\d{2}/\d{2,4}\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})',
                # Pattern 9: Flexible date at start - capture everything and parse amounts
                r'^(\d{2}/\d{2}/\d{2,4})\s+(.+)$'
            ]
            
            transaction_found = False
            
            # Try to match current line with patterns
            for pattern_num, pattern in enumerate(hdfc_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    try:
                        raw_date = match.group(1)
                        date_str = self._convert_hdfc_date(raw_date)
                        
                        if not date_str or date_str == raw_date:
                            logger.warning(f"HDFC date conversion failed for '{raw_date}', skipping")
                            continue
                        
                        # Handle different pattern structures
                        if pattern_num in [1]:  # Full format with reference number (Date Desc RefNo Date Amount Balance)
                            description = match.group(2).strip()
                            amount_str = match.group(4).replace(',', '').replace('-', '')
                            balance_str = match.group(5)
                                
                        elif pattern_num in [2, 3]:  # Simple format DD/MM/YYYY or DD/MM/YY (Date Desc Amount Balance)
                            description = match.group(2).strip()
                            amount_str = match.group(3).replace(',', '').replace('-', '')
                            balance_str = match.group(4)
                            
                        elif pattern_num in [4]:  # With reference number (Date Desc RefNo Amount Balance)
                            description = match.group(2).strip()
                            # Remove reference number from description
                            description = re.sub(r'\d{8,}', '', description).strip()
                            amount_str = match.group(4).replace(',', '').replace('-', '')
                            balance_str = match.group(5)
                            
                        elif pattern_num in [5, 6]:  # HDFC with ref number and negative amounts (Date Desc RefNo -Amount Balance)
                            description = match.group(2).strip()
                            # Remove reference number from description if it exists
                            description = re.sub(r'\d{8,}', '', description).strip()
                            amount_str = match.group(4).replace(',', '').replace('-', '')
                            balance_str = match.group(5)
                            
                        elif pattern_num == 7:  # Date + Description + negative amount + balance (Date Desc -Amount Balance)
                            description = match.group(2).strip()
                            amount_str = match.group(3).replace(',', '').replace('-', '')
                            balance_str = match.group(4)
                            
                        elif pattern_num == 8:  # ATM/EAW transactions (Date EAW-/ATW-/ATM- Desc RefNo Date Amount Balance)
                            prefix = match.group(2).strip()  # EAW-, ATW-, ATM-
                            description = prefix + match.group(3).strip()
                            # Remove reference number from description if it exists
                            description = re.sub(r'\d{8,}', '', description).strip()
                            amount_str = match.group(5).replace(',', '').replace('-', '')
                            balance_str = match.group(6)
                            
                        elif pattern_num == 9:  # Flexible - extract amounts from line
                            full_content = match.group(2).strip()
                            amounts = re.findall(r'(-?[\d,]+\.\d{2})', full_content)
                            
                            if len(amounts) >= 1:
                                # Build description by removing amounts
                                description = full_content
                                for amt in amounts:
                                    description = description.replace(amt, ' ')
                                description = re.sub(r'\d{8,}', '', description)  # Remove ref numbers
                                description = re.sub(r'\s+', ' ', description).strip()
                                
                                # If we have multiple amounts, first is usually transaction amount, last is balance
                                if len(amounts) >= 2:
                                    # Find the transaction amount (look for negative or smaller positive amount)
                                    trans_amounts = [amt for amt in amounts if '-' in amt or float(amt.replace(',', '').replace('-', '')) < 50000]
                                    if trans_amounts:
                                        amount_str = trans_amounts[0].replace(',', '').replace('-', '')
                                    else:
                                        amount_str = amounts[0].replace(',', '').replace('-', '')
                                    balance_str = amounts[-1]  # Last amount is usually balance
                                else:
                                    amount_str = amounts[0].replace(',', '').replace('-', '')
                                    balance_str = None
                            else:
                                continue
                        
                        # Clean and validate description
                        description = re.sub(r'\s+', ' ', description).strip()
                        if not description or len(description) < 3:
                            description = 'HDFC Transaction'
                        
                        # **ENHANCED HDFC COLUMN-BASED CLASSIFICATION**
                        # HDFC Format: Date | Narration | Chq./Ref.No. | Value Dt | Withdrawal Amt. | Deposit Amt. | Closing Balance
                        # Key insight: The column position determines the transaction type!
                        
                        desc_lower = description.lower()
                        line_lower = line.lower()
                        
                        # Parse all amounts from the line to understand column structure
                        all_amounts = re.findall(r'([\d,]+\.\d{2})', line)
                        
                        # Determine transaction type based on HDFC column analysis and context
                        trans_type = self._determine_hdfc_transaction_type_by_columns(
                            description, amount_str, balance_str, all_amounts, line
                        )
                        
                        # Validate amount
                        amount_str = amount_str.replace(',', '') if amount_str else '0'
                        
                        if float(amount_str) > 0:
                            logger.info(f"*** HDFC TRANSACTION FOUND ON LINE {line_num} (Pattern {pattern_num}) ***")
                            logger.info(f"  Date: {date_str}")
                            logger.info(f"  Description: {description}")
                            logger.info(f"  Amount: {amount_str}")
                            logger.info(f"  Type: {trans_type}")
                            
                            transactions.append({
                                'date_str': date_str,
                                'description': description,
                                'amount_str': amount_str,
                                'type': trans_type,
                                'source_line': line,
                                'pattern_used': pattern_num,
                                'bank_type': 'HDFC'
                            })
                            
                            transaction_found = True
                            break
                            
                    except Exception as e:
                        logger.error(f"Error parsing HDFC transaction on line {line_num}: {str(e)}")
                        continue
            
            # Enhanced fallback parsing - catch any missed transactions
            if not transaction_found:
                # Look for any line with date pattern and amounts
                date_match = re.search(r'(\d{2}/\d{2}/\d{2,4})', line)
                amounts = re.findall(r'([\d,]+\.\d{2})', line)
                
                if date_match and len(amounts) >= 1:
                    try:
                        raw_date = date_match.group(1)
                        date_str = self._convert_hdfc_date(raw_date)
                        
                        if date_str and date_str != raw_date:
                            # Extract description by removing date and amounts
                            description = line
                            description = re.sub(r'\d{2}/\d{2}/\d{2,4}', '', description)
                            for amt in amounts:
                                description = description.replace(amt, ' ')
                            description = re.sub(r'\d{8,}', '', description)  # Remove reference numbers
                            description = re.sub(r'\s+', ' ', description).strip()
                            
                            if not description or len(description) < 3:
                                description = 'HDFC Transaction'
                            
                            amount_str = amounts[0].replace(',', '')
                            
                            # Use the same column-based classification for fallback transactions
                            trans_type = self._determine_hdfc_transaction_type_by_columns(
                                description, amount_str, amounts[-1] if len(amounts) > 1 else None, amounts, line
                            )
                            
                            if float(amount_str) > 0:
                                logger.info(f"*** HDFC FALLBACK TRANSACTION ON LINE {line_num} ***")
                                logger.info(f"  Date: {date_str}")
                                logger.info(f"  Description: {description}")
                                logger.info(f"  Amount: {amount_str}")
                                logger.info(f"  Type: {trans_type}")
                                
                                transactions.append({
                                    'date_str': date_str,
                                    'description': description,
                                    'amount_str': amount_str,
                                    'type': trans_type,
                                    'source_line': line,
                                    'pattern_used': 'fallback',
                                    'bank_type': 'HDFC'
                                })
                            
                    except Exception as e:
                        logger.error(f"Error in HDFC fallback parsing on line {line_num}: {str(e)}")
            
            i += 1

        logger.info(f"=== FOUND {len(transactions)} HDFC TRANSACTIONS ===")
        return transactions

    def _determine_hdfc_transaction_type_by_columns(self, description: str, amount_str: str, balance_str: str, all_amounts: list, full_line: str) -> str:
        """
        Determine HDFC transaction type using column-based analysis
        
        HDFC Format: Date | Narration | Chq./Ref.No. | Value Dt | Withdrawal Amt. | Deposit Amt. | Closing Balance
        
        Key Logic:
        - Withdrawal Amt. column = expense (money going out)
        - Deposit Amt. column = income (money coming in) 
        - ALL UPI- transactions are outgoing payments (expenses) regardless of recipient
        """
        
        desc_lower = description.lower()
        line_lower = full_line.lower()
        
        # **CRITICAL UPI RULE**: ALL UPI- transactions are expenses (outgoing payments)
        # This overrides all other logic because UPI- prefix indicates money leaving account
        if desc_lower.startswith('upi-'):
            return 'expense'
        
        # Strong income indicators (override column analysis)
        strong_income_indicators = [
            'interest paid', 'salary', 'wage', 'dividend', 'bonus', 'refund',
            'cashback', 'commission', 'reversal', 'credit interest'
        ]
        
        if any(indicator in desc_lower for indicator in strong_income_indicators):
            return 'income'
        
        # Company payment detection (legitimate income sources)
        company_patterns = [
            r'\b(software|tech|technologies|solutions|systems|consulting|services)\s+(p\s*l|pvt\s*ltd)',
            r'\b[a-z0-9]+\s*(software|tech|solutions|systems|services)\b',
            r'\b(microsoft|google|amazon|apple|oracle|sap)\b(?!.*\b(recharge|payment|purchase)\b)'
        ]
        
        for pattern in company_patterns:
            if re.search(pattern, desc_lower):
                return 'income'
        
        # Strong expense indicators
        strong_expense_indicators = [
            'atw-', 'atm-', 'eaw-', 'pos ', 'nwd-', 'withdrawal', 'purchase',
            'bill payment', 'recharge', 'fee', 'charge', 'emi', 'loan'
        ]
        
        if any(indicator in desc_lower for indicator in strong_expense_indicators):
            return 'expense'
        
        # Merchant/Service payments (always expense)
        merchant_patterns = [
            r'\b(zomato|swiggy|uber|ola|rapido|amazon|flipkart|myntra|nykaa)\b',
            r'\b(airtel|jio|vodafone)\b.*\brecharge\b',
            r'\bgoogle.*\brecharge\b'
        ]
        
        for pattern in merchant_patterns:
            if re.search(pattern, desc_lower):
                return 'expense'
        
        # Column-based analysis using balance change
        try:
            transaction_amount = float(amount_str.replace(',', ''))
            
            if balance_str:
                closing_balance = float(balance_str.replace(',', ''))
                
                # Large amounts analysis
                if transaction_amount >= 5000:
                    # Large amounts are typically income unless clearly expense
                    if any(term in desc_lower for term in ['payment', 'purchase', 'bill']):
                        return 'expense'
                    else:
                        return 'income'
                
                # Medium amounts (1000-5000)
                elif transaction_amount >= 1000:
                    # Context-based decision
                    if any(term in desc_lower for term in ['store', 'mart', 'shop', 'restaurant']):
                        return 'expense'
                    else:
                        return 'income'  # Could be person-to-person payment received
                
                # Small amounts (< 1000) - usually expenses
                else:
                    return 'expense'
                    
        except (ValueError, TypeError):
            pass
        
        # Default fallback based on amount size
        try:
            amount_val = float(amount_str.replace(',', ''))
            return 'income' if amount_val >= 2000 else 'expense'
        except:
            return 'expense'

    def _parse_axis_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse AXIS Bank specific format."""
        transactions = []
        
        logger.info(f"=== AXIS PARSING ===")
        
        for i, line in enumerate(lines):
            line = line.strip()
            line_num = i + 1
            
            if len(line) < 8:
                continue

            # Skip AXIS header/footer lines
            skip_indicators = [
                'axis bank', 'prakhya venkata', 'joint holder', 'customer id',
                'ifsc code', 'micr code', 'nominee', 'scheme', 'tran date',
                'particulars', 'debit', 'credit', 'balance', 'legends',
                'transaction total', 'registered office', 'this is a system'
            ]
            
            if any(skip in line.lower() for skip in skip_indicators):
                continue

            logger.info(f"AXIS Line {line_num}: {line}")

            # AXIS patterns
            axis_patterns = [
                # Pattern 1: DD-MM-YYYY Description Amount Balance Init
                r'^(\d{2}-\d{2}-\d{4})(.+?)\s+([\d,]+\.00)\s+([\d,]+\.00)\s+\d+$',
                # Pattern 2: UPI/IMPS/RTGS transactions
                r'^(\d{2}-\d{2}-\d{4})(UPI|IMPS|RTGS).+?\s+([\d,]+\.00)\s+([\d,]+\.00)\s+\d+$',
                # Pattern 3: ATM transactions
                r'^(\d{2}-\d{2}-\d{4})(ATM-CASH).+?\s+([\d,]+\.00)\s+([\d,]+\.00)\s+\d+$'
            ]
            
            transaction_found = False
            
            for pattern_num, pattern in enumerate(axis_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    try:
                        raw_date = match.group(1)
                        date_str = self._convert_axis_date(raw_date)
                        
                        if not date_str:
                            continue
                        
                        description = match.group(2).strip()
                        amount_str = match.group(3).replace(',', '')
                        balance_str = match.group(4)
                        
                        # Determine transaction type
                        desc_lower = description.lower()
                        if any(term in desc_lower for term in ['credit', 'deposit', 'received']):
                            trans_type = 'income'
                        else:
                            trans_type = 'expense'
                        
                        if float(amount_str) > 0:
                            logger.info(f"*** AXIS TRANSACTION FOUND ON LINE {line_num} (Pattern {pattern_num}) ***")
                            logger.info(f"  Date: {date_str}")
                            logger.info(f"  Description: {description}")
                            logger.info(f"  Amount: {amount_str}")
                            logger.info(f"  Type: {trans_type}")
                            
                            transactions.append({
                                'date_str': date_str,
                                'description': description,
                                'amount_str': amount_str,
                                'type': trans_type,
                                'source_line': line,
                                'pattern_used': pattern_num,
                                'bank_type': 'AXIS'
                            })
                            
                            transaction_found = True
                            break
                            
                    except Exception as e:
                        logger.error(f"Error parsing AXIS transaction on line {line_num}: {str(e)}")
                        continue
        
        logger.info(f"=== FOUND {len(transactions)} AXIS TRANSACTIONS ===")
        return transactions

    def _parse_federal_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse Federal Bank specific format."""
        transactions = []
        
        logger.info(f"=== FEDERAL PARSING ===")
        
        for i, line in enumerate(lines):
            line = line.strip()
            line_num = i + 1
            
            if len(line) < 8:
                continue

            # Skip Federal header/footer lines
            skip_indicators = [
                'federal bank', 'corporate office', 'prakhya venkata', 'branch name',
                'customer id', 'swift code', 'currency', 'date value date',
                'particulars', 'withdrawals', 'deposits', 'balance',
                'abbreviations used', 'grand total'
            ]
            
            if any(skip in line.lower() for skip in skip_indicators):
                continue

            logger.info(f"FEDERAL Line {line_num}: {line}")

            # Federal patterns
            federal_patterns = [
                # Pattern 1: DD-MMM-YYYY DD-MMM-YYYY Description TFR SXXXXXXXX Amount Balance Cr/Dr
                r'^(\d{2}-[A-Z]{3}-\d{4})\s+(\d{2}-[A-Z]{3}-\d{4})\s+(.+?)TFR\s+S\d+\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(Cr|Dr)$',
                # Pattern 2: UPI transactions
                r'^(\d{2}-[A-Z]{3}-\d{4})\s+(\d{2}-[A-Z]{3}-\d{4})\s+(UPI\s+IN|UPI\s*OUT).+?TFR\s+S\d+\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(Cr|Dr)$'
            ]
            
            transaction_found = False
            
            for pattern_num, pattern in enumerate(federal_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    try:
                        raw_date = match.group(1)
                        date_str = self._convert_federal_date(raw_date)
                        
                        if not date_str:
                            continue
                        
                        if pattern_num == 1:
                            description = match.group(3).strip()
                            amount_str = match.group(4).replace(',', '')
                            balance_str = match.group(5)
                            cr_dr = match.group(6)
                        else:
                            description = match.group(3).strip()
                            amount_str = match.group(4).replace(',', '')
                            balance_str = match.group(5)
                            cr_dr = match.group(6)
                        
                        # Determine transaction type based on Cr/Dr
                        trans_type = 'income' if cr_dr.lower() == 'cr' else 'expense'
                        
                        if float(amount_str) > 0:
                            logger.info(f"*** FEDERAL TRANSACTION FOUND ON LINE {line_num} (Pattern {pattern_num}) ***")
                            logger.info(f"  Date: {date_str}")
                            logger.info(f"  Description: {description}")
                            logger.info(f"  Amount: {amount_str}")
                            logger.info(f"  Type: {trans_type}")
                            
                            transactions.append({
                                'date_str': date_str,
                                'description': description,
                                'amount_str': amount_str,
                                'type': trans_type,
                                'source_line': line,
                                'pattern_used': pattern_num,
                                'bank_type': 'FEDERAL'
                            })
                            
                            transaction_found = True
                            break
                            
                    except Exception as e:
                        logger.error(f"Error parsing FEDERAL transaction on line {line_num}: {str(e)}")
                        continue
        
        logger.info(f"=== FOUND {len(transactions)} FEDERAL TRANSACTIONS ===")
        return transactions

    def _convert_axis_date(self, date_str: str) -> str:
        """Convert AXIS date format DD-MM-YYYY to YYYY-MM-DD."""
        try:
            parts = date_str.split('-')
            if len(parts) == 3:
                day, month, year = parts
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            return None
        except:
            return None

    def _convert_federal_date(self, date_str: str) -> str:
        """Convert Federal date format DD-MMM-YYYY to YYYY-MM-DD."""
        try:
            month_map = {
                'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
                'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
            }
            
            parts = date_str.split('-')
            if len(parts) == 3:
                day, month_abbr, year = parts
                month = month_map.get(month_abbr.upper())
                if month:
                    return f"{year}-{month}-{day.zfill(2)}"
            return None
        except:
            return None

    def _parse_generic_bank_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Generic bank statement parsing for unknown formats."""
        transactions = []
        
        logger.info(f"=== GENERIC BANK PARSING ===")
        
        # Generic patterns that might work across different banks
        generic_patterns = [
            # Pattern for date + description + amount + Dr/Cr
            r'^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+(.+?)\s+([\d,]+\.\d{2})\s+(Dr|Cr|DR|CR)',
            # Pattern for transactions with balance
            r'^(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(Dr|Cr|DR|CR)'
        ]
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if len(line) < 15:
                continue

            for pattern_num, pattern in enumerate(generic_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    try:
                        if pattern_num == 1:  # Date-based pattern
                            date_str = self._normalize_date(match.group(1))
                            description = match.group(2).strip()
                            amount_str = match.group(3)
                            dr_cr = match.group(4).upper()
                        else:  # Amount-based pattern
                            description = match.group(1).strip()
                            amount_str = match.group(2)
                            dr_cr = match.group(4).upper()
                            date_str = statement_date or datetime.now().strftime('%d/%m/%Y')

                        trans_type = 'expense' if dr_cr in ['DR', 'DEBIT'] else 'income'
                        
                        transactions.append({
                            'date_str': date_str,
                            'description': description,
                            'amount_str': amount_str,
                            'type': trans_type,
                            'source_line': line,
                            'pattern_used': pattern_num,
                            'bank_type': 'GENERIC'
                        })
                        break
                        
                    except Exception as e:
                        logger.error(f"Error in generic parsing: {str(e)}")
                        continue

        logger.info(f"=== FOUND {len(transactions)} GENERIC TRANSACTIONS ===")
        return transactions

    def _extract_federal_bank_dates(self, pdf_text: str) -> List[str]:
        """Extract dates in Federal Bank format (DD-MMM-YYYY)."""
        dates = []
        
        # Federal Bank uses DD-MMM-YYYY format
        federal_date_pattern = r'\b(\d{1,2}-[A-Z]{3}-\d{4})\b'
        matches = re.findall(federal_date_pattern, pdf_text)
        
        for match in matches:
            converted_date = self._convert_federal_date(match)
            if converted_date and converted_date not in dates:
                dates.append(converted_date)
        
        return dates

    def _convert_federal_date(self, federal_date: str) -> str:
        """Convert Federal Bank date format (22-MAY-2023) to DD/MM/YYYY."""
        try:
            # Parse DD-MMM-YYYY and convert to DD/MM/YYYY
            date_obj = datetime.strptime(federal_date, '%d-%b-%Y')
            return date_obj.strftime('%d/%m/%Y')
        except ValueError:
            logger.error(f"Failed to parse Federal Bank date: {federal_date}")
            return federal_date

    def _convert_sbi_date(self, date_str: str) -> str:
        """Convert SBI date formats to DD/MM/YYYY."""
        try:
            # Try SBI format: "01-08-23" (DD-MM-YY)
            if re.match(r'\d{2}-\d{2}-\d{2}', date_str):
                date_obj = datetime.strptime(date_str, '%d-%m-%y')
                return date_obj.strftime('%d/%m/%Y')
            
            # Try SBI format: "01 Aug 2023" (DD MMM YYYY)
            elif re.match(r'\d{1,2} [A-Za-z]{3} \d{4}', date_str):
                date_obj = datetime.strptime(date_str, '%d %b %Y')
                return date_obj.strftime('%d/%m/%Y')
            
            # Try SBI format: "01 JUN 2024" (DD MMM YYYY uppercase)
            elif re.match(r'\d{1,2} [A-Z]{3} \d{4}', date_str):
                date_obj = datetime.strptime(date_str, '%d %B %Y')
                return date_obj.strftime('%d/%m/%Y')
                
            else:
                # Fallback to original logic
                date_obj = datetime.strptime(date_str, '%d-%m-%y')
                return date_obj.strftime('%d/%m/%Y')
                
        except ValueError as e:
            logger.error(f"Error converting SBI date '{date_str}': {e}")
            return date_str

    def _convert_sbi_date_format2(self, date_str: str) -> str:
        """Convert SBI date format (DD MMM YYYY) to YYYY-MM-DD with enhanced error handling."""
        if not date_str:
            logger.warning("Empty date string provided to SBI date converter")
            return None
            
        date_str = date_str.strip()
        
        try:
            # SBI format from PDF: "01 JUN 2024"
            dt_obj = datetime.strptime(date_str, '%d %b %Y')
            formatted_date = dt_obj.strftime('%Y-%m-%d')
            logger.debug(f"SBI date conversion: '{date_str}' -> '{formatted_date}'")
            return formatted_date
        except ValueError as e:
            logger.error(f"Error converting SBI date format2 '{date_str}': {e}")
            # Try alternative SBI formats
            try:
                # Try DD-MMM-YYYY format
                dt_obj = datetime.strptime(date_str, '%d-%b-%Y')
                return dt_obj.strftime('%Y-%m-%d')
            except ValueError:
                logger.error(f"Could not parse SBI date in any known format: '{date_str}'")
                return date_str

    def _convert_hdfc_date(self, date_str: str) -> str:
        """Convert HDFC date format to YYYY-MM-DD with enhanced format handling."""
        if not date_str:
            logger.warning("Empty date string provided to HDFC date converter")
            return None
            
        date_str = date_str.strip()
        logger.debug(f"Converting HDFC date: '{date_str}'")
        
        try:
            # Handle DD/MM/YYYY format (like "01/06/2024")
            if '/' in date_str and len(date_str.split('/')[2]) == 4:
                dt_obj = datetime.strptime(date_str, "%d/%m/%Y")
                formatted_date = dt_obj.strftime("%Y-%m-%d")
                logger.debug(f"HDFC 4-digit year: '{date_str}' -> '{formatted_date}'")
                return formatted_date
            
            # Handle DD/MM/YY format (like "01/06/24") - most common HDFC format
            elif '/' in date_str and len(date_str.split('/')[2]) == 2:
                dt_obj = datetime.strptime(date_str, '%d/%m/%y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"HDFC 2-digit year: '{date_str}' -> '{formatted_date}'")
                return formatted_date
            
            # Handle DD-MMM-YYYY format (like "01-JUN-2024") 
            elif '-' in date_str and len(date_str.split('-')) == 3:
                dt_obj = datetime.strptime(date_str, "%d-%b-%Y")
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"HDFC month name format: '{date_str}' -> '{formatted_date}'")
                return formatted_date
                
            # Handle DD MMM YYYY format (like "01 JUN 2024")
            elif ' ' in date_str and len(date_str.split(' ')) == 3:
                dt_obj = datetime.strptime(date_str, "%d %b %Y")
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"HDFC space-separated format: '{date_str}' -> '{formatted_date}'")
                return formatted_date
                
        except ValueError as e:
            logger.error(f"Error parsing HDFC date '{date_str}': {str(e)}")
        except Exception as e:
            logger.error(f"Unexpected error parsing HDFC date '{date_str}': {str(e)}")
            
        # If all parsing fails, return None to indicate failure
        logger.warning(f"Could not parse HDFC date format: '{date_str}', all formats failed")
        return None

    def _parse_axis_transactions(self, lines: List[str], statement_date: str = None) -> List[Dict]:
        """Parse Axis Bank specific format."""
        transactions = []
        
        logger.info(f"=== AXIS BANK PARSING ===")
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if len(line) < 15:
                continue

            # Skip Axis Bank header/footer lines
            skip_indicators = [
                'axis bank', 'statement of account', 'account number:', 'joint holder', 'customer id',
                'customer name:', 'branch:', 'ifsc code:', 'micr code:', 'registered mobile',
                'registered email', 'scheme :', 'ckyc number', 'nominee', 'opening balance',
                'closing balance', 'statement summary', 'transaction codes', 'end of statement',
                'request from:', 'this is a system generated', 'please contact the branch'
            ]
            
            if any(skip in line.lower() for skip in skip_indicators):
                continue

            logger.info(f"Axis Line {line_num}: {line}")

            # Axis Bank format patterns
            # Format: DD-MM-YYYYDESCRIPTION                          AMOUNT             BALANCE BRANCH
            axis_patterns = [
                # Main pattern: DD-MM-YYYY + Description + Amount + Balance + Branch
                r'^(\d{2}-\d{2}-\d{4})(.+?)\s+(\d[\d,]*\.\d{2})\s+(\d[\d,]*\.\d{2})\s+(\d+)\s*$',
                
                # Alternative pattern with debit amount at end
                r'^(\d{2}-\d{2}-\d{4})(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+\d+\s*$'
            ]
            
            transaction_found = False
            
            for pattern_num, pattern in enumerate(axis_patterns, 1):
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    try:
                        # Both patterns have: date, description, amount, balance, branch
                        date_str = self._convert_axis_date(match.group(1))
                        description = match.group(2).strip()
                        amount_str = match.group(3).replace(',', '')
                        balance = match.group(4)
                        
                        # Determine transaction type based on description keywords
                        trans_type = self._classify_axis_transaction(description)
                        
                        if float(amount_str) > 0:
                            logger.info(f"*** AXIS TRANSACTION FOUND ON LINE {line_num} ***")
                            logger.info(f"  Date: {date_str}")
                            logger.info(f"  Description: {description}")
                            logger.info(f"  Amount: {amount_str}")
                            logger.info(f"  Type: {trans_type}")
                            
                            transactions.append({
                                'date_str': date_str,
                                'description': description,
                                'amount_str': amount_str,
                                'type': trans_type,
                                'source_line': line,
                                'pattern_used': pattern_num,
                                'bank_type': 'AXIS'
                            })
                            
                            transaction_found = True
                            break
                            
                    except Exception as e:
                        logger.error(f"Error parsing Axis transaction on line {line_num}: {str(e)}")
                        continue
        
        logger.info(f"=== FOUND {len(transactions)} AXIS TRANSACTIONS ===")
        return transactions

    def _convert_axis_date(self, axis_date: str) -> str:
        """Convert Axis Bank date formats to YYYY-MM-DD with enhanced error handling."""
        if not axis_date:
            logger.warning("Empty date string provided to Axis date converter")
            return None
            
        axis_date = axis_date.strip()
        
        try:
            # Handle DD-MMM-YY format (01-Aug-23) - case insensitive
            if re.match(r'\d{1,2}-[A-Za-z]{3}-\d{2}', axis_date):
                dt_obj = datetime.strptime(axis_date, '%d-%b-%y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"Axis date conversion (DD-MMM-YY): '{axis_date}' -> '{formatted_date}'")
                return formatted_date
                
            # Handle DD/MM/YY format  
            elif re.match(r'\d{1,2}/\d{1,2}/\d{2}', axis_date):
                dt_obj = datetime.strptime(axis_date, '%d/%m/%y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"Axis date conversion (DD/MM/YY): '{axis_date}' -> '{formatted_date}'")
                return formatted_date
                
            # Handle DD-MM-YYYY format
            elif re.match(r'\d{1,2}-\d{1,2}-\d{4}', axis_date):
                dt_obj = datetime.strptime(axis_date, '%d-%m-%Y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"Axis date conversion (DD-MM-YYYY): '{axis_date}' -> '{formatted_date}'")
                return formatted_date
                
            # Handle DD/MM/YYYY format
            elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', axis_date):
                dt_obj = datetime.strptime(axis_date, '%d/%m/%Y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"Axis date conversion (DD/MM/YYYY): '{axis_date}' -> '{formatted_date}'")
                return formatted_date
                
            # Handle DD-MMM-YYYY format (01-Aug-2024)
            elif re.match(r'\d{1,2}-[A-Za-z]{3}-\d{4}', axis_date):
                dt_obj = datetime.strptime(axis_date, '%d-%b-%Y')
                formatted_date = dt_obj.strftime('%Y-%m-%d')
                logger.debug(f"Axis date conversion (DD-MMM-YYYY): '{axis_date}' -> '{formatted_date}'")
                return formatted_date
                
            # Fallback for unknown formats
            else:
                logger.warning(f"Unknown Axis date format: '{axis_date}', attempting flexible parsing")
                # Try a few more common formats
                fallback_formats = ['%d %b %Y', '%Y-%m-%d', '%m/%d/%Y']
                for fmt in fallback_formats:
                    try:
                        dt_obj = datetime.strptime(axis_date, fmt)
                        formatted_date = dt_obj.strftime('%Y-%m-%d')
                        logger.debug(f"Axis fallback date conversion: '{axis_date}' -> '{formatted_date}'")
                        return formatted_date
                    except ValueError:
                        continue
                        
                logger.error(f"Could not parse Axis date in any known format: '{axis_date}'")
                return axis_date
                
        except Exception as e:
            logger.error(f"Error converting Axis date '{axis_date}': {str(e)}")
            return axis_date

    def _classify_axis_transaction(self, description: str) -> str:
        """Classify Axis Bank transaction as income or expense based on description."""
        desc_lower = description.lower()
        
        # Income indicators
        income_keywords = [
            'upi/p2a', 'imps/p2a', 'rtgs', 'neft', 'salary', 'interest', 
            'dividend', 'refund', 'cashback', 'bonus', 'deposit', 
            'credit', 'received', 'transfer from'
        ]
        
        # Expense indicators
        expense_keywords = [
            'upi/p2m', 'atm-cash', 'pos', 'purchase', 'withdrawal', 
            'debit', 'payment', 'bill', 'emi', 'loan', 'fee', 'charge',
            'swiggy', 'zomato', 'amazon', 'flipkart', 'google', 'cinema'
        ]
        
        # Check for income patterns first
        for keyword in income_keywords:
            if keyword in desc_lower:
                return 'income'
                
        # Check for expense patterns
        for keyword in expense_keywords:
            if keyword in desc_lower:
                return 'expense'
                
        # Default classification based on common patterns
        if any(word in desc_lower for word in ['cash', 'withdrawal', 'purchase']):
            return 'expense'
        elif any(word in desc_lower for word in ['deposit', 'credit', 'received']):
            return 'income'
            
        # Default to expense if unclear
        return 'expense'

    def _process_federal_bank_transaction(self, match, current_date: str, current_description: str, 
                                        previous_balance: float, line_num: int) -> Dict:
        """Process a matched Federal Bank transaction."""
        try:
            reference_part = match.group(1).strip()
            amount_str = match.group(2).strip()
            balance_str = match.group(3).strip()
            cr_dr_indicator = match.group(4).strip()
            extra_part = match.group(5).strip() if len(match.groups()) > 4 else ""

            # Convert amounts to float for calculation
            transaction_amount = float(amount_str.replace(',', ''))
            new_balance = float(balance_str.replace(',', ''))

            # Build full description
            full_description = current_description
            if reference_part:
                full_description += " " + reference_part
            if extra_part:
                full_description += " " + extra_part

            # Determine transaction type based on balance change
            if previous_balance == 0:  # First transaction
                actual_trans_type = 'income'
            else:
                balance_change = new_balance - previous_balance
                actual_trans_type = 'income' if balance_change > 0 else 'expense'

            # Override with description-based detection for accuracy
            desc_based_type = self._determine_federal_bank_transaction_type(full_description, cr_dr_indicator)
            
            # Use description-based type if it's more specific
            if desc_based_type == actual_trans_type or 'upi' in full_description.lower():
                final_type = desc_based_type
            else:
                final_type = actual_trans_type

            logger.info(f"*** FEDERAL BANK TRANSACTION PROCESSED ON LINE {line_num} ***")
            logger.info(f"  Date: {current_date}")
            logger.info(f"  Description: {full_description[:50]}...")
            logger.info(f"  Amount: {amount_str}")
            logger.info(f"  Type: {final_type}")

            return {
                'date_str': current_date,
                'description': full_description,
                'amount_str': amount_str,
                'type': final_type,
                'source_line': match.group(0),
                'pattern_used': 1,
                'balance_change': new_balance - previous_balance,
                'new_balance': new_balance,
                'bank_type': 'FEDERAL'
            }
            
        except Exception as e:
            logger.error(f"Error processing Federal Bank transaction: {str(e)}")
            return None

    def _normalize_date(self, date_str: str) -> str:
        """Normalize various date formats to DD/MM/YYYY."""
        try:
            # Remove any extra whitespace
            date_str = date_str.strip()
            
            # Try different date formats
            formats = [
                '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',  # DD/MM/YYYY variants
                '%d/%m/%y', '%d-%m-%y', '%d.%m.%y',  # DD/MM/YY variants  
                '%Y-%m-%d', '%Y/%m/%d',              # YYYY-MM-DD variants
                '%m/%d/%Y', '%m-%d-%Y',              # MM/DD/YYYY variants
                '%d %b %Y', '%d-%b-%Y',              # DD MMM YYYY variants
            ]
            
            for fmt in formats:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    return date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    continue
                    
            # If no format matches, return original
            logger.warning(f"Could not normalize date format: {date_str}")
            return date_str
            
        except Exception as e:
            logger.error(f"Error normalizing date '{date_str}': {e}")
            return date_str

    def _determine_federal_bank_transaction_type(self, description: str, cr_dr: str) -> str:
        """Determine transaction type specifically for Federal Bank format."""
        description_lower = description.lower()
        
        logger.info(f"Analyzing Federal Bank transaction: '{description}'")
        
        # Federal Bank specific patterns from your statement
        
        # INCOME indicators (money coming in)
        income_indicators = [
            'upi in',           # UPI IN = money received
            'upi credit',       # UPI credit
            'salary',           # Salary payments
            'epifi technologies',  # Your employer
            'credit',           # General credits
            'deposit',          # Deposits
            'refund',           # Refunds
            'cashback',         # Cashback
            'dividend',         # Investment returns
            'interest',         # Interest credits
            'transfer from',    # Money transferred to you
            'received from',    # Money received
        ]
        
        # EXPENSE indicators (money going out)
        expense_indicators = [
            'upi out',          # UPI OUT = money paid
            'upiout',           # UPI OUT (no space)
            'payment to',       # Payments made
            'paytm',            # Paytm payments
            'bill payment',     # Bill payments
            'withdrawal',       # Cash withdrawals
            'charge',           # Bank charges
            'fee',              # Various fees
            'purchase',         # Purchases
            'shopping',         # Shopping
            'fuel',             # Fuel payments
            'grocery',          # Grocery payments
            'restaurant',       # Food payments
            'transfer to',      # Money sent to others
        ]
        
        # Check for strong income indicators
        for indicator in income_indicators:
            if indicator in description_lower:
                logger.info(f"  >>> FEDERAL BANK INCOME: '{indicator}' found <<<")
                return 'income'
        
        # Check for strong expense indicators
        for indicator in expense_indicators:
            if indicator in description_lower:
                logger.info(f"  >>> FEDERAL BANK EXPENSE: '{indicator}' found <<<")
                return 'expense'
        
        # For Federal Bank, if no clear indicator, analyze the description context
        # UPI transactions with specific patterns
        if 'upi' in description_lower:
            if any(pattern in description_lower for pattern in ['@', 'qr', 'pay']):
                logger.info(f"  >>> UPI PAYMENT PATTERN - EXPENSE <<<")
                return 'expense'
            else:
                logger.info(f"  >>> UPI GENERIC - fallback to Cr/Dr <<<")
        
        # Technology/company names usually indicate income
        if any(company in description_lower for company in ['tech', 'technologies', 'pvt', 'ltd']):
            logger.info(f"  >>> COMPANY PAYMENT - INCOME <<<")
            return 'income'
        
        # Final fallback to Cr/Dr (but Federal Bank format may need special handling)
        if cr_dr.upper() == 'CR':
            logger.info(f"  >>> FEDERAL BANK CREDIT - INCOME (fallback) <<<")
            return 'income'
        else:
            logger.info(f"  >>> FEDERAL BANK DEBIT - EXPENSE (fallback) <<<")
            return 'expense'

    def _extract_transaction_dates(self, pdf_text: str, statement_date: str) -> List[str]:
        """Extract all potential transaction dates from PDF."""
        dates = []
        
        # Look for date patterns in the text
        date_patterns = [
            r'\b(\d{1,2}/\d{1,2}/\d{4})\b',
            r'\b(\d{4}-\d{1,2}-\d{1,2})\b',
            r'\b(\d{1,2}-\d{1,2}-\d{4})\b'
        ]
        
        for pattern in date_patterns:
            matches = re.findall(pattern, pdf_text)
            for match in matches:
                if match not in dates:
                    # Validate date format
                    try:
                        if '/' in match:
                            datetime.strptime(match, '%d/%m/%Y')
                        elif '-' in match and len(match.split('-')[0]) == 4:
                            datetime.strptime(match, '%Y-%m-%d')
                        else:
                            datetime.strptime(match, '%d-%m-%Y')
                        dates.append(match)
                    except ValueError:
                        continue
        
        # Add statement date if provided
        if statement_date and statement_date not in dates:
            dates.append(statement_date)
            
        return dates

    def _get_transaction_date(self, line: str, available_dates: List[str], statement_date: str) -> str:
        """Get the most appropriate transaction date for a line."""
        # First, check if the line itself contains a date
        for date_str in available_dates:
            if date_str in line:
                return date_str
        
        # Fall back to statement date or first available date
        if statement_date:
            return statement_date
        elif available_dates:
            return available_dates[0]
        else:
            # Final fallback to current date
            return datetime.now().strftime('%d/%m/%Y')

    def _determine_transaction_type_enhanced(self, description: str, cr_dr: str, amount_str: str) -> str:
        """Enhanced logic to determine if transaction is expense or income."""
        description_lower = description.lower()
        
        # Enhanced keyword lists based on real banking patterns
        strong_expense_indicators = [
            # Direct payments and purchases
            'payment to', 'paid to', 'purchase', 'shopping', 'retail', 
            'grocery', 'supermarket', 'restaurant', 'cafe', 'hotel',
            # Bills and utilities  
            'electricity bill', 'water bill', 'gas bill', 'phone bill', 'internet bill',
            'credit card payment', 'loan payment', 'emi payment', 'insurance premium',
            # Withdrawals and fees
            'atm withdrawal', 'cash withdrawal', 'bank charges', 'service charge',
            'annual fee', 'processing fee', 'late payment', 'penalty',
            # Transport and fuel
            'petrol', 'diesel', 'fuel', 'gas station', 'uber', 'taxi', 'ola',
            'bus fare', 'train ticket', 'flight booking', 'parking fee',
            # Medical and education
            'hospital', 'medical', 'pharmacy', 'doctor', 'clinic', 'medicine',
            'school fee', 'college fee', 'education', 'course fee', 'exam fee',
            # Transfers out
            'transfer to', 'sent to', 'remittance', 'wire transfer'
        ]
        
        strong_income_indicators = [
            # Salary and employment
            'salary credit', 'salary deposit', 'wage credit', 'payroll', 'bonus credit',
            'overtime payment', 'commission credit', 'incentive credit',
            # Business income
            'revenue credit', 'sales credit', 'invoice payment', 'client payment',
            'freelance payment', 'consulting fee', 'service income',
            # Investment returns
            'dividend credit', 'interest credit', 'fd interest', 'rd interest',
            'mutual fund', 'stock dividend', 'capital gain', 'investment return',
            # Refunds and returns
            'refund credit', 'cashback', 'reward points', 'loyalty bonus',
            'insurance claim', 'tax refund', 'gst refund',
            # Transfers in
            'transfer from', 'received from', 'deposit from', 'credit transfer',
            'family transfer', 'gift received'
        ]
        
        # Check for strong indicators first
        for indicator in strong_expense_indicators:
            if indicator in description_lower:
                logger.info(f"  >>> STRONG EXPENSE INDICATOR: '{indicator}' <<<")
                return 'expense'
                
        for indicator in strong_income_indicators:
            if indicator in description_lower:
                logger.info(f"  >>> STRONG INCOME INDICATOR: '{indicator}' <<<")
                return 'income'
        
        # Check for general patterns
        general_expense_patterns = [
            'payment', 'purchase', 'bill', 'fee', 'charge', 'withdrawal',
            'debit', 'spent', 'bought', 'paid'
        ]
        
        general_income_patterns = [
            'credit', 'deposit', 'received', 'earned', 'bonus', 'salary',
            'refund', 'cashback', 'dividend', 'interest'
        ]
        
        expense_score = sum(1 for pattern in general_expense_patterns if pattern in description_lower)
        income_score = sum(1 for pattern in general_income_patterns if pattern in description_lower)
        
        if expense_score > income_score:
            logger.info(f"  >>> EXPENSE by pattern score: {expense_score} vs {income_score} <<<")
            return 'expense'
        elif income_score > expense_score:
            logger.info(f"  >>> INCOME by pattern score: {income_score} vs {expense_score} <<<")
            return 'income'
        
        # Final fallback to Cr/Dr with correction
        # Many banks show all transactions as Credit in statements, so be careful
        if cr_dr.upper() == 'DR':
            logger.info(f"  >>> EXPENSE by DR indicator (fallback) <<<")
            return 'expense'
        else:
            # For Credit entries, default to income but this might need user review
            logger.info(f"  >>> INCOME by CR indicator (fallback - review recommended) <<<")
            return 'income'

    def _flexible_pdf_parsing(self, pdf_text: str, statement_date: str = None) -> List[Dict]:
        """Fallback parsing method for PDFs that don't match standard patterns."""
        transactions = []
        lines = pdf_text.split('\n')

        # Look for any line with a date-like pattern and an amount
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if len(line) < 15:
                continue

            # Find date patterns
            date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line)
            if not date_match:
                continue

            # Find amount patterns
            amount_matches = re.findall(r'(\d{1,3}(?:,\d{3})*\.\d{2})', line)
            if not amount_matches:
                continue

            # Extract the most likely amount (usually the last one)
            amount_str = amount_matches[-1]
            date_str = date_match.group(1)

            # Extract description (everything between date and amount)
            desc_start = date_match.end()
            amount_start = line.rfind(amount_str)

            if amount_start > desc_start:
                description = line[desc_start:amount_start].strip()
                description = re.sub(r'\s+', ' ', description)

                if len(description) > 3:
                    trans_type = self._determine_transaction_type(line, description, amount_str)

                    transactions.append({
                        'date_str': date_str,
                        'description': description,
                        'amount_str': amount_str,
                        'type': trans_type,
                        'source_line': line
                    })

        logger.info(f"Flexible parsing found {len(transactions)} potential transactions")
        return transactions

    def _determine_transaction_type(self, line: str, description: str, amount_str: str) -> str:
        """Determine if transaction is income or expense based on context."""
        line_lower = line.lower()
        desc_lower = description.lower()

        # **CRITICAL UPI LOGIC**: ALL UPI- transactions in bank statements are outgoing payments (expenses)
        # Income via UPI would show as "UPI CREDIT" or "RECEIVED FROM" not "UPI-"
        if desc_lower.startswith('upi-'):
            return 'expense'  # All UPI- transactions are outgoing payments

        # UPI Credits/Receipts (actual income via UPI)
        if any(term in desc_lower for term in [
            'upi credit', 'received from', 'credit from', 'transfer from',
            'payment received', 'money received'
        ]):
            return 'income'

        # Company/Professional income indicators (high priority)
        company_indicators = [
            'software', 'solutions', 'services', 'technologies', 'systems',
            'consulting', 'pvt ltd', 'private limited', 'ltd', 'inc',
            'corporation', 'corp', 'company', 'co', 'salary', 'wage'
        ]
        
        if any(indicator in desc_lower for indicator in company_indicators):
            return 'income'

        # Strong income indicators
        income_keywords = [
            'deposit', 'salary', 'wage', 'payment received', 'credit', 'transfer in',
            'interest', 'dividend', 'refund', 'cashback', 'bonus'
        ]

        # Strong expense indicators
        expense_keywords = [
            'withdrawal', 'purchase', 'payment', 'debit', 'fee', 'charge',
            'atm', 'atw', 'eaw', 'transfer out', 'check', 'automatic payment'
        ]

        # Check for explicit debit/credit indicators
        if re.search(r'\b(cr|credit|\+)\b', line_lower):
            return 'income'
        elif re.search(r'\b(dr|debit|\-)\b', line_lower):
            return 'expense'

        # Check description for keywords
        for keyword in income_keywords:
            if keyword in desc_lower:
                return 'income'

        for keyword in expense_keywords:
            if keyword in desc_lower:
                return 'expense'

        # Default to expense if uncertain (conservative approach)
        return 'expense'

    def _create_pdf_transaction(
        self,
        trans_data: Dict,
        row_num: int,
        account: Account,
        user,  # User model instance
        family_group=None
    ) -> Optional[Dict]:
        """Create transaction object from parsed PDF data."""
        try:
            # Parse date
            transaction_date = self._parse_date(trans_data['date_str'])
            if not transaction_date:
                raise ValueError(f"Invalid date format: {trans_data['date_str']}")

            # Parse amount
            amount_str = trans_data['amount_str'].replace(',', '').replace('$', '').replace('₹', '')
            try:
                amount = Decimal(amount_str)
                if amount <= 0:
                    raise ValueError("Amount must be positive")
            except (InvalidOperation, ValueError):
                raise ValueError(f"Invalid amount format: {amount_str}")

            # Clean description
            description = trans_data['description'][:200]  # Limit length
            if not description:
                raise ValueError("Description is required")

            # Get transaction type
            trans_type = trans_data['type']
            if trans_type not in ['income', 'expense']:
                trans_type = 'expense'  # Default fallback

            return {
                'amount': amount,
                'description': description,
                'transaction_type': trans_type,
                'category': None,  # PDF parsing doesn't typically include categories
                'account': account,
                'date': transaction_date,
                'user': user,
                'family_group': family_group,
                'imported_from': 'pdf_upload',
                'notes': f'Imported from PDF statement'
            }

        except Exception as e:
            raise ValueError(f"Failed to parse transaction: {str(e)}")


class TransactionExportService:
    """Service for exporting transactions to various formats."""

    def export_transactions_csv(
        self,
        transactions,
        filename: str = "transactions_export.csv"
    ) -> Tuple[str, str]:
        """Export transactions to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Date',
            'Description',
            'Amount',
            'Type',
            'Category',
            'Account',
            'Notes'
        ])

        # Write transactions
        for transaction in transactions:
            writer.writerow([
                transaction.date.strftime('%Y-%m-%d'),
                transaction.description,
                str(transaction.amount),
                transaction.get_transaction_type_display(),
                transaction.category.name if transaction.category else '',
                transaction.account.name if transaction.account else '',
                transaction.notes or ''
            ])

        return output.getvalue(), filename

    def export_transactions_excel(
        self,
        transactions,
        filename: str = "transactions_export.xlsx"
    ) -> Tuple[bytes, str]:
        """Export transactions to Excel format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Write header
        headers = [
            'Date', 'Description', 'Amount', 'Type',
            'Category', 'Account', 'Notes'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write transactions
        for row_num, transaction in enumerate(transactions, 2):
            ws.cell(row=row_num, column=1, value=transaction.date)
            ws.cell(row=row_num, column=2, value=transaction.description)
            ws.cell(row=row_num, column=3, value=float(transaction.amount))
            ws.cell(row=row_num, column=4, value=transaction.get_transaction_type_display())
            ws.cell(row=row_num, column=5, value=transaction.category.name if transaction.category else '')
            ws.cell(row=row_num, column=6, value=transaction.account.name if transaction.account else '')
            ws.cell(row=row_num, column=7, value=transaction.notes or '')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read(), filename


# Service instances
import_service = TransactionImportService()
export_service = TransactionExportService()